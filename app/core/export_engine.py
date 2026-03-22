"""
출력 엔진 (Excel)

체크리스트 데이터를 Excel(.xlsx) 파일로 생성합니다.
MVP-1에서는 Excel 출력만 구현합니다.
"""

import logging
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from app.core.checklist_builder import ChecklistBuilder
from app.config.logger import setup_logger
from app.config.settings import (
    EXCEL_AREA_SEP_COLOR,
    EXCEL_FILL_COL_COLOR,
    EXCEL_HEADER_COLOR,
    EXCEL_HEADER_FONT_COLOR,
    EXCEL_RESULT_COL_COLOR,
    EXCEL_ROW_HEIGHT,
)
from app.models.schemas import PMItem, PMPeriod

logger = setup_logger(__name__)

# 스타일 상수
HEADER_FILL = PatternFill(start_color="1A3A5C", end_color="1A3A5C", fill_type="solid")
HEADER_FONT = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="맑은 고딕", size=9)
TITLE_FONT = Font(name="맑은 고딕", size=14, bold=True, color="1A3A5C")
SUBTITLE_FONT = Font(name="맑은 고딕", size=11, bold=True, color="333333")
INFO_FONT = Font(name="맑은 고딕", size=10, color="666666")

CHECK_FILL = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")  # 연한 노랑
PERSON_FILL = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")  # 연한 파랑
AREA_SEPARATOR_FILL = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")  # 영역 구분

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# 컬럼 설정
COLUMNS = [
    ("No.", 6),
    ("영역", 15),
    ("부위", 18),
    ("점검 항목", 35),
    ("점검 방법", 18),
    ("기준값", 20),
    ("점검 결과", 12),
    ("비고", 20),
    ("담당자", 12),
    ("점검일", 14),
]

PERIOD_NAMES = {
    PMPeriod.DAILY: "일간",
    PMPeriod.MONTHLY: "월간",
    PMPeriod.QUARTERLY: "분기",
    PMPeriod.SEMI_ANNUAL: "반기",
    PMPeriod.ANNUAL: "연간",
}


class ExportEngine:
    """체크리스트 출력 엔진"""

    def __init__(self):
        self.builder = ChecklistBuilder()

    def generate_excel(
        self,
        manual_name: str,
        equipment_name: str,
        all_items: list[PMItem],
    ) -> bytes:
        """
        Excel 체크리스트 생성

        Args:
            manual_name: 매뉴얼 파일명
            equipment_name: 설비명
            all_items: 전체 PM 항목 리스트

        Returns:
            xlsx 파일 바이트 데이터
        """
        wb = Workbook()

        from app.models.schemas import clean_text
        safe_manual = clean_text(manual_name) or "매뉴얼"
        safe_equip = clean_text(equipment_name) or "설비"

        # 1. 요약 시트
        self._create_summary_sheet(wb, safe_manual, safe_equip, all_items)

        # 2. 주기별 PM 시트
        by_period = self.builder.build_by_period(all_items)
        for period in PMPeriod:
            period_items = by_period.get(period.value, [])
            sheet_name = f"{PERIOD_NAMES[period]} PM"
            self._create_pm_sheet(wb, sheet_name, period_items, safe_equip)

        # 3. 전체 목록 시트
        self._create_pm_sheet(wb, "전체 목록", all_items, safe_equip)

        # BytesIO로 저장
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        logger.info(f"Excel 파일 생성 완료: {len(all_items)}개 항목")
        return output.getvalue()

    def _create_summary_sheet(
        self,
        wb: Workbook,
        manual_name: str,
        equipment_name: str,
        items: list[PMItem],
    ):
        """요약 시트 생성"""
        ws = wb.active
        ws.title = "요약"

        # 제목
        ws.merge_cells("A1:J1")
        ws["A1"] = "🔧 PM 체크리스트 요약"
        ws["A1"].font = TITLE_FONT
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40

        # 설비 정보
        info_data = [
            ("설비명", equipment_name),
            ("매뉴얼", manual_name),
            ("생성일", datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("총 PM 항목 수", str(len(items))),
        ]

        row = 3
        for label, value in info_data:
            ws[f"B{row}"] = label
            ws[f"B{row}"].font = Font(name="맑은 고딕", size=10, bold=True)
            ws[f"C{row}"] = value
            ws[f"C{row}"].font = BODY_FONT
            row += 1

        # 통계 테이블
        stats = self.builder.get_statistics(items)

        row += 1
        ws[f"B{row}"] = "주기별 항목 수"
        ws[f"B{row}"].font = SUBTITLE_FONT
        row += 1

        # 통계 헤더
        for col, period_name in enumerate(["일간", "월간", "분기", "반기", "연간"], start=2):
            cell = ws.cell(row=row, column=col, value=period_name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

        row += 1
        period_map = {"일": "일", "월": "월", "분기": "분기", "반기": "반기", "년": "년"}
        for col, period_key in enumerate(["일", "월", "분기", "반기", "년"], start=2):
            count = stats["by_period"].get(period_key, 0)
            cell = ws.cell(row=row, column=col, value=count)
            cell.font = BODY_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

        # 영역별 항목 수
        row += 2
        ws[f"B{row}"] = "영역별 항목 수"
        ws[f"B{row}"].font = SUBTITLE_FONT
        row += 1

        for area_name, count in sorted(stats["by_area"].items()):
            ws[f"B{row}"] = area_name
            ws[f"B{row}"].font = BODY_FONT
            ws[f"C{row}"] = count
            ws[f"C{row}"].font = BODY_FONT
            ws[f"B{row}"].border = THIN_BORDER
            ws[f"C{row}"].border = THIN_BORDER
            row += 1

        # 부위 목록
        row += 1
        ws[f"B{row}"] = f"부위 수: {stats['total_parts']}개"
        ws[f"B{row}"].font = SUBTITLE_FONT
        row += 1
        ws[f"B{row}"] = ", ".join(stats["parts_list"])
        ws[f"B{row}"].font = INFO_FONT

        # 컬럼 너비 조정
        ws.column_dimensions["A"].width = 4
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 40

    def _create_pm_sheet(
        self,
        wb: Workbook,
        sheet_name: str,
        items: list[PMItem],
        equipment_name: str,
    ):
        """주기별/전체 PM 시트 생성"""
        ws = wb.create_sheet(title=sheet_name)

        # 제목 행
        ws.merge_cells("A1:J1")
        ws["A1"] = f"📋 {equipment_name} — {sheet_name} 체크리스트"
        ws["A1"].font = TITLE_FONT
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 35

        # 항목이 없는 경우
        if not items:
            ws["A3"] = "해당 주기의 PM 항목이 없습니다."
            ws["A3"].font = INFO_FONT
            return

        # 헤더 행 (행 3)
        header_row = 3
        for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=col_name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER
            ws.column_dimensions[get_column_letter(col_idx)].width = col_width

        ws.row_dimensions[header_row].height = 25

        # 데이터 행
        current_area = ""
        data_row = header_row + 1

        for idx, item in enumerate(items, start=1):
            # 영역 변경 시 구분선 삽입
            if item.area != current_area:
                current_area = item.area
                ws.merge_cells(f"A{data_row}:J{data_row}")
                area_cell = ws.cell(row=data_row, column=1, value=f"▸ {current_area}")
                area_cell.font = Font(name="맑은 고딕", size=9, bold=True, color="1A3A5C")
                area_cell.fill = AREA_SEPARATOR_FILL
                area_cell.border = THIN_BORDER
                ws.row_dimensions[data_row].height = 22
                data_row += 1

            # 데이터 셀
            row_data = [
                idx,
                item.area,
                item.equipment_part,
                item.item_name,
                item.method,
                item.standard_value or "-",
                "",  # 점검 결과
                item.note or "",
                "",  # 담당자
                "",  # 점검일
            ]

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=data_row, column=col_idx, value=value)
                cell.font = BODY_FONT
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = THIN_BORDER

                # 점검 결과 열: 연한 노란색
                if col_idx == 7:
                    cell.fill = CHECK_FILL
                # 담당자, 점검일 열: 연한 파란색
                elif col_idx in (9, 10):
                    cell.fill = PERSON_FILL

            ws.row_dimensions[data_row].height = 22
            data_row += 1

        # 마지막 행에 합계 정보
        data_row += 1
        ws[f"A{data_row}"] = f"총 {len(items)}개 항목"
        ws[f"A{data_row}"].font = INFO_FONT

    def save_to_file(
        self,
        manual_name: str,
        equipment_name: str,
        all_items: list[PMItem],
        output_path: str,
    ) -> str:
        """
        Excel 파일을 지정 경로에 저장

        Args:
            manual_name: 매뉴얼 파일명
            equipment_name: 설비명
            all_items: PM 항목 리스트
            output_path: 저장할 파일 경로

        Returns:
            저장된 파일 경로
        """
        data = self.generate_excel(manual_name, equipment_name, all_items)

        with open(output_path, "wb") as f:
            f.write(data)

        logger.info(f"Excel 파일 저장 완료: {output_path}")
        return output_path
